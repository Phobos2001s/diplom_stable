import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


# https://habr.com/ru/companies/otus/articles/331998/


class Excel:
    EXCEL_PATH = 'files/template.xlsx'
    OUT_PATH = 'files/generated/'
    START_ROW = 10
    START_COL = 2


    def load(self, path):
        # Load in the workbook
        if '.xlsx' in path:
            wb = load_workbook(path)
        else:
            wb = self.cvt_xls_to_xlsx(path)

        return wb

    def save(self, wb, file_name):
        if '.xlsx' not in file_name:
            file_name += '.xlsx'
        wb.save(self.OUT_PATH + file_name)

    # все листы в датафреймы
    def read_df(self, wb):
        # Get sheet names
        ##print(wb.get_sheet_names())
        data_frame = []
        for sheet_name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_name)
            print(sheet.title)
            data_frame.append(self.ws_to_df(sheet))

        ##print(data_frame)
        return data_frame


    def read_df1(self, wb):
        # Get sheet names
        ##print(wb.get_sheet_names())
        data_frame = []
        for sheet_name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_name)
            print(sheet.title)
            data_frame.append(self.ws_to_df1(sheet))

        ##print(data_frame)
        return data_frame

    # весь лист в датафрейм
    def ws_to_df(self, sheet):
        # Convert Sheet to DataFrame
        # df = pd.DataFrame(sheet.values)

        # Put the sheet values in `data`
        data = [[sheet.cell(r, c).value for c in range(self.START_COL, sheet.max_column + 1)] for r in
                range(self.START_ROW, sheet.max_row + 1)]

        # Convert your data to a list
        data = list(data)

        # Make your DataFrame
        df = pd.DataFrame(data)
        df.fillna(0, inplace=True)
        return df

    def ws_to_df1(self, sheet):
        # Convert Sheet to DataFrame
        # df = pd.DataFrame(sheet.values)

        # Put the sheet values in `data`
        data = [[sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)] for r in
                range(self.START_ROW, sheet.max_row + 1)]

        # Convert your data to a list
        data = list(data)

        # Make your DataFrame
        df = pd.DataFrame(data)
        df.fillna(0, inplace=True)
        return df

    def cvt_xls_to_xlsx(self, src_file_path):
        book_xls = xlrd.open_workbook(src_file_path)
        book_xlsx = Workbook()
        sheet_names = book_xls.sheet_names()
        for sheet_index, sheet_name in enumerate(sheet_names):
            sheet_xls = book_xls.sheet_by_name(sheet_name)
            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_name
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)
        return book_xlsx

    # запись из датафрейма непосредственно в страницу эксель файла
    def df_to_wb(self, dataframe, worksheet):
        index = self.START_ROW + 1
        max_col = worksheet.max_column - self.START_COL + 1
        for r in dataframe_to_rows(dataframe, index=False, header=False):
            for col in range(0, max_col):
                # print(r[col])
                worksheet.cell(row=index, column=self.START_COL + col).value = r[col]
            index += 1
